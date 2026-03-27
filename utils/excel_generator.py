import os
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from models import Quotation, CompanyConfig

def generate_excel(quotation_id, lang='cn'):
    """生成Excel报价单"""
    quotation = Quotation.query.get_or_404(quotation_id)
    company = CompanyConfig.query.first()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Quotation' if lang == 'en' else '报价单'

    # 样式定义
    title_font = Font(name='Arial', size=16, bold=True, color='0066CC')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    light_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

    # 语言相关文本
    if lang == 'en':
        texts = {
            'quotation': 'QUOTATION',
            'date': 'Date',
            'valid_until': 'Valid Until',
            'bill_to': 'Bill To',
            'no': 'No.',
            'image': 'Image',
            'description': 'Description',
            'specification': 'Specification',
            'unit_price': 'Unit Price',
            'quantity': 'Qty',
            'unit': 'Unit',
            'amount': 'Amount',
            'weight': 'Weight(kg)',
            'volume': 'Volume(m³)',
            'total': 'TOTAL',
            'total_weight': 'Total Weight',
            'total_volume': 'Total Volume',
            'remarks': 'Remarks',
            'bank_info': 'Bank Information'
        }
    else:
        texts = {
            'quotation': '报价单',
            'date': '日期',
            'valid_until': '有效期至',
            'bill_to': '买方信息',
            'no': '序号',
            'image': '图片',
            'description': '商品名称',
            'specification': '规格',
            'unit_price': '单价',
            'quantity': '数量',
            'unit': '单位',
            'amount': '金额',
            'weight': '重量(kg)',
            'volume': '体积(m³)',
            'total': '合计',
            'total_weight': '总重量',
            'total_volume': '总体积',
            'remarks': '备注',
            'bank_info': '银行信息'
        }

    row = 1

    # 公司名称
    company_name = company.name_en if lang == 'en' and company.name_en else company.name_cn
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = company_name
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].alignment = center_align
    ws.row_dimensions[row].height = 30
    row += 1

    # 公司联系信息
    company_address = company.address_en if lang == 'en' and company.address_en else company.address_cn
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = f"{company_address} | Tel: {company.phone} | Email: {company.email}"
    ws[f'A{row}'].font = Font(name='Arial', size=9, color='666666')
    ws[f'A{row}'].alignment = center_align
    row += 2

    # 币种符号
    currency_symbol = '$' if quotation.currency == 'USD' else '¥'

    # 报价单信息
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = f"{texts['quotation']}: {quotation.quotation_no}"
    ws[f'A{row}'].font = bold_font

    ws.merge_cells(f'F{row}:J{row}')
    ws[f'F{row}'] = f"{texts['date']}: {quotation.date}  {texts['valid_until']}: {quotation.valid_until or '-'}"
    ws[f'F{row}'].font = normal_font
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 2

    # 买方信息
    customer = quotation.customer
    if customer:
        name = customer.name_en if lang == 'en' and customer.name_en else customer.name_cn
        phone = customer.phone or '-'
        email = customer.email or '-'
        address = customer.address or '-'
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = f"{texts['bill_to']}: {name}  Tel: {phone}  Email: {email}  Address: {address}"
        ws[f'A{row}'].font = Font(name='Arial', size=9)
        ws[f'A{row}'].fill = light_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'A{row}'].border = thin_border
    row += 2

    # 表头
    headers = [texts['no'], texts['image'], texts['description'], texts['specification'],
               texts['unit_price'], texts['quantity'], texts['unit'], texts['amount'],
               texts['weight'], texts['volume']]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    ws.row_dimensions[row].height = 25
    row += 1

    # 数据行
    for idx, item in enumerate(quotation.items, 1):
        # 检查产品是否存在
        if item.product is None:
            continue

        name = item.product.name_en if lang == 'en' and item.product.name_en else item.product.name_cn
        spec = item.product.specification_en if lang == 'en' and item.product.specification_en else item.product.specification_cn
        unit = item.product.unit_en if lang == 'en' and item.product.unit_en else item.product.unit_cn

        data = [
            idx,
            '',  # 图片列 - 稍后插入
            name,
            spec or '-',
            f"{currency_symbol}{item.price:.2f}",
            item.quantity,
            unit,
            f"{currency_symbol}{item.subtotal:.2f}",
            f"{item.weight:.2f}",
            f"{item.volume:.3f}"
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
            if col in [1, 6, 7]:
                cell.alignment = center_align
            elif col in [5, 8, 9, 10]:
                cell.alignment = right_align
            else:
                cell.alignment = left_align

        # 插入图片
        if item.product.image_path:
            img_path = os.path.join(os.getcwd(), 'static', 'uploads', item.product.image_path)
            if os.path.exists(img_path):
                try:
                    img = XLImage(img_path)
                    # 设置图片大小
                    img.width = 60
                    img.height = 60
                    # 设置行高
                    ws.row_dimensions[row].height = 50
                    # 插入图片到B列
                    ws.add_image(img, f'B{row}')
                except Exception as e:
                    print(f"Error inserting image: {e}")

        row += 1

    # 合计行
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = texts['total']
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'A{row}'].fill = light_fill
    ws[f'A{row}'].border = thin_border

    ws.cell(row=row, column=8, value=f"{currency_symbol}{quotation.total_price:.2f}").font = bold_font
    ws.cell(row=row, column=8).alignment = right_align
    ws.cell(row=row, column=8).fill = light_fill
    ws.cell(row=row, column=8).border = thin_border

    ws.cell(row=row, column=9, value=f"{quotation.total_weight:.2f}").font = bold_font
    ws.cell(row=row, column=9).alignment = right_align
    ws.cell(row=row, column=9).fill = light_fill
    ws.cell(row=row, column=9).border = thin_border

    ws.cell(row=row, column=10, value=f"{quotation.total_volume:.3f}").font = bold_font
    ws.cell(row=row, column=10).alignment = right_align
    ws.cell(row=row, column=10).fill = light_fill
    ws.cell(row=row, column=10).border = thin_border

    row += 2

    # 备注
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = f"{texts['remarks']}: {quotation.remarks or '-'}"
    ws[f'A{row}'].font = normal_font
    row += 2

    # 银行信息
    bank_name = company.bank_name_en if lang == 'en' and company.bank_name_en else company.bank_name_cn
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = f"{texts['bank_info']}: {bank_name} | Account: {company.bank_account} | SWIFT: {company.bank_code}"
    ws[f'A{row}'].font = Font(name='Arial', size=9)
    row += 1

    payment_note = company.payment_note_en if lang == 'en' and company.payment_note_en else company.payment_note_cn
    if payment_note:
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = f"Payment Note: {payment_note}"
        ws[f'A{row}'].font = Font(name='Arial', size=9, italic=True)

    # 设置列宽
    column_widths = [6, 12, 25, 15, 15, 8, 8, 15, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 保存文件
    excel_path = os.path.join(os.getcwd(), 'static', 'uploads', f'quotation_{quotation_id}.xlsx')
    wb.save(excel_path)

    return send_file(excel_path, as_attachment=True, download_name=f'{quotation.quotation_no}.xlsx')